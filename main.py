

import random
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from PIL import Image,ImageTk
import pyperclip
from openpyxl import load_workbook
import openpyxl
import pathlib
import numpy as np
import smtplib
import socket
from email.message import EmailMessage
import re 

file1=pathlib.Path('Resources\Saved_Passwords.xlsx')

if file1.exists():
    pass
else:
    file1=openpyxl.Workbook()
    sheet=file1.active
    sheet["A1"]="Account name"
    sheet["B1"]="Password"
    sheet["C1"]="Email address"

    file1.save('Resources\Saved_Passwords.xlsx')

characters="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890!@#$%^&*()_+}{"


regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
 
value=False     

def check(email):

    if(re.search(regex,email)):  
        return True          
    else:  
        return False  


def verify(otp,otp_entry,email,frame):
    otp_entry=otp_entry.get()
    frame.destroy()
    if otp_entry==otp:

        file1=openpyxl.load_workbook('Resources\Saved_Passwords.xlsx')
        sheet=file1.active
        if sheet["C2"].value==None:
            sheet=file1.active
            sheet["C2"]=email
            file1.save('Resources\Saved_Passwords.xlsx')
            main()
        else:
            saved_pass(frame)


    else:
        messagebox.showerror("One Time Password Generator","Invalid OTP")
        otp_screen(otp,email)
    

def otp_screen(otp,email):

    frame=tk.Frame(window,width=650,height=420)
    frame.place(x=0,y=0)

    lab0=tk.Label(frame,image =bg_image)
    lab0.place(x=-2,y=-2)

    lab1=tk.Label(frame,text="OTP has been sent to registered email address !!",bg="#66B2FF",fg="black",font=("verdana",15))
    lab1.place(x=40,y=50)

    otp_lab=tk.Label(frame,text="Enter OTP : ",bg="#66B2FF",fg="black",font=("verdana",13))
    otp_lab.place(x=150,y=200)

    otp_entry=tk.Entry(frame,bg='white',fg='black',width=20,font=("verdana",12),borderwidth=2,justify='left',relief='groove')
    otp_entry.place(x=260,y=200)

    submit_but=tk.Button(frame,image=submit_img,bg='#66B2FF',width=130,height=30,relief='flat',command=lambda:verify(otp,otp_entry,email,frame))
    submit_but.place(x=260,y=270)


def sign_up(frame0,email_entry):
    file1=openpyxl.load_workbook("Resources\Saved_Passwords.xlsx")
    sheet=file1.active
    if sheet["C2"].value!=None:
        email=sheet["C2"].value
    else:
        email=email_entry.get()
    if check(email):
        otp=''
        for i in range(6):
            otp+=str(random.randint(0,9))

        otp_send(email,otp,frame0)
        otp_screen(otp,email)
        
    else:
        messagebox.showerror("One Time Password Generator","Invalid Email address")
        registration(window)

def registration(window):
    
    frame0=tk.Frame(window,width=650,height=420)
    frame0.place(x=0,y=0)

    bg_label=tk.Label(frame0,image=bg_image)
    bg_label.place(x=-3,y=0)

    login_bg=tk.Label(frame0,image =login_bg_image)
    login_bg.place(x=120,y=70)

    user=tk.Label(frame0,image =user_img)
    user.place(x=275,y=10)

    sign_up_text=tk.Label(frame0,text="Sign Up",bg="#D5D5D5",fg="blue",font=("Georgia",40))
    sign_up_text.place(x=230,y=90)

    email_lab=tk.Label(frame0,text="Enter email address:",bg="#D5D5D5",fg="black",font=("verdana",13))
    email_lab.place(x=140,y=190)

    email_img_lab=tk.Label(frame0,image =email_img)
    email_img_lab.place(x=140,y=220)

    email_entry=tk.Entry(frame0,bg='white',fg='black',width=32,font=("verdana",12),borderwidth=2,justify='left',relief='groove')
    email_entry.place(x=180,y=220)

    sign_up_but=tk.Button(frame0,image=sign_up_img,bg='white',width=130,height=35,relief='flat',command=lambda:sign_up(frame0,email_entry))
    sign_up_but.place(x=250,y=280)

    window.mainloop()


def otp_send(to,otp,frame2):
    try :
        body="To verify your email address, please use the following One Time Password (OTP): {} \nDo not share this OTP with anyone".format(otp)
        msg=EmailMessage()
        msg.set_content(body)
        msg['subject']="OTP Verification"
        msg['to']=to

        user='official.randompass.generator@gmail.com'
        msg['from']=user
        password="dicnibhtohntatct"

        server=smtplib.SMTP("smtp.gmail.com",587)
        server.starttls()
        server.login(user,password)
        server.send_message(msg)
        server.quit()
    
    except socket.gaierror:
        messagebox.showerror("One Time Password Generator","Internal Server Error")
        back(frame2)
        

def encr(text):
    encrypt_text=''
    for i in text:
        pos=characters.find(i)
        new_pos=(pos+10)%76
        encrypt_text+=characters[new_pos]
    return encrypt_text

def decr(text):
    file1=openpyxl.load_workbook("Resources\Saved_Passwords.xlsx")
    sheet=file1.active
    if sheet.cell(row=2,column=1).value==None:
        messagebox.showinfo("One Time Password Generator","No data Found")

    else:
        decrypt_text=''
        for i in text:
            pos=characters.find(i)
            new_pos=(pos-10)%76
            decrypt_text+=characters[new_pos]
        return decrypt_text

def password_saved_window():
    frame4=tk.Frame(window,width=650,height=420,bg='white')
    frame4.place(x=0,y=80)

    back_but=tk.Button(frame4,image=back_img,bg='white',fg='white',relief='flat',justify='center',command=lambda:back(frame4))
    back_but.place(x=0,y=0)

    save_succ=tk.Label(frame4,text="Password Saved Successfully !!!",bg='white',font=('veradana bold',20))
    save_succ.place(x=110,y=120)

    saved_pass_but=tk.Button(frame4,text="View Saved Passwords",width=20,height=2,fg="dark blue",bg="light blue",borderwidth=2,relief='ridge',justify='center',activeforeground="dark blue",font=("verdana Bold",10),command=lambda:sign_up(frame4,email_entry=tk.Entry()))
    saved_pass_but.place(x=210,y=200)

def save_pass(password,account_name):
    if account_name.get()=="":
        messagebox.showerror("One Time Password Generator","Account name is mendatory")

    else:
        file1=load_workbook("Resources\Saved_Passwords.xlsx")
        sheet=file1.active
        sheet.cell(column=1,row=sheet.max_row,value=account_name.get())
        sheet.cell(column=2,row=sheet.max_row,value=encr(password))
        file1.save("Resources\Saved_Passwords.xlsx")

        password_saved_window()

def copy(treev):
    select=treev.focus()
    saved_password=treev.item(select).get("values")[2]
    pyperclip.copy(saved_password)

def delete(treev,frame1):
    select=treev.focus()
    del_row=saved_password=treev.item(select).get("values")[0]
    del_row+=1

    file = "Resources\Saved_Passwords.xlsx"
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    
    if del_row==ws.max_row:
        ws["A{}".format(del_row)].value = None
        ws["B{}".format(del_row)].value = None

    else:
        ws.move_range("A{}:B{}".format(del_row+1,ws.max_row), rows=-1)

    wb.save(file)
    messagebox.showinfo("One Time Password Generator","Account Information Deleted Successfully !!")
    saved_pass(frame1)


def saved_pass(frame1):

    frame1.destroy()
    frame3=tk.Frame(window,width=650,height=420,bg='white')
    frame3.place(x=0,y=80)

    back_but=tk.Button(frame3,image=back_img,bg='white',fg='white',relief='flat',justify='center',command=lambda:back(frame3))
    back_but.place(x=0,y=0)

    file = "Resources\Saved_Passwords.xlsx"
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    data_list=np.array([[0]*2]*(ws.max_row-1),dtype='object')
    '''for x in range (2,ws.max_row):
        data_list.append([])'''
    
    for x in range(1,ws.max_row):
        data_list[x-1][0]=ws.cell(row=x+1,column=1).value
        data_list[x-1][1]=decr(ws.cell(row=x+1,column=2).value)
        

    style=ttk.Style()
    style.theme_use('clam')
    style.configure("Treeview",
                bg='silver',
                fg='black',
                rowheight=25,
                fieldbackground='white'
                )
    style.map("Treeview",background=[('selected','light green')],foreground=[('selected','red')])
    treev = ttk.Treeview(frame3,show='headings', selectmode ='browse',height=12) 
    treev.place(x=70,y=5)

    verscrlbar = ttk.Scrollbar(frame3,
                            orient ="vertical",  
                            command = treev.yview,) 
    
    verscrlbar.place(x=625,y=6,height=322)
    
    treev.configure(xscrollcommand = verscrlbar.set,) 
    treev["columns"] = ("1", "2","3")
    
    treev.column("1", width = 50, anchor ='c')
    treev.column("2", width = 200, anchor ='w')
    treev.column("3", width = 280, anchor ='c')
    
    treev.heading("1",text="Sr. No.")
    treev.heading("2", text ="Account Name") 
    treev.heading("3", text ="Password") 
    
    for x in range(0,ws.max_row-1):
        treev.insert("",'end',values=(x+1,data_list[x][0],data_list[x][1]))

    m =tk.Menu(frame3, tearoff = 0) 
    m.add_command(label ="Copy",command=lambda:copy(treev))
    m.add_command(label ="Delete",command=lambda:delete(treev,frame1)) 
    m.add_separator()

    

    def do_popup(event): 
        try: 
            m.tk_popup(event.x_root, event.y_root) 
        finally: 
            m.grab_release() 
    
    treev.bind("<Button-3>", do_popup) 

def back(frame2):
    
    frame2.destroy()
    main()
    
def display(window,password,frame1):

    frame2=tk.Frame(window,width=650,height=420,bg='white')
    frame2.place(x=0,y=80)

    back_but=tk.Button(frame2,image=back_img,bg='white',fg='white',relief='flat',justify='center',command=lambda:back(frame2))
    back_but.place(x=0,y=0)

    pass_text=tk.Label(frame2,text="Generated password is : ",bg='white',font=('veradana bold',15))
    pass_text.place(x=92,y=100)

    value=tk.StringVar()
    pass_entry=tk.Entry(frame2,textvariable=value,borderwidth=3,relief="groove",font=('veradana bold',15))
    pass_entry.place(x=350,y=100)

    value.set(password)

    m =tk.Menu(frame2, tearoff = 0) 
    m.add_command(label ="Copy",command=lambda:pyperclip.copy(password)) 
    m.add_separator()

    pass_entry.config(state='disabled',bg='white')  


    account_text=tk.Label(frame2,text="Name of account : ",bg='white',font=('veradana bold',15))
    account_text.place(x=150,y=200)

    value=tk.StringVar()
    account_entry=tk.Entry(frame2,textvariable=value,borderwidth=3,relief="groove",font=('veradana bold',15))
    account_entry.place(x=350,y=200)

    
    def do_popup(event): 
        try: 
            m.tk_popup(event.x_root, event.y_root) 
        finally: 
            m.grab_release() 
    
    pass_entry.bind("<Button-3>", do_popup) 
    

    save_but=tk.Button(frame2,image=save_img,relief='flat',bg='white',fg='white',activeforeground='white',activebackground='white',justify='center',font=("verdana Bold",12),command=lambda:save_pass(password,account_entry))
    save_but.place(x=250,y=265)


def generate(frame1,length_list,window):
    pass_char="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890!@#$%^&*()_+}{"
    length=int(length_list.get())
    
    password=''

    for i in range(length):
        password+=random.choice(pass_char)
    
    frame1.destroy()
    display(window,password,frame1)

def main():
    
    text_lab=tk.Label(window,text="       One Time Password Generator",height=2,width=31,fg='blue',bg='light yellow',font=('veradana bold',27))
    text_lab.place(x=0,y=0)

    lock_img=Image.open("Resources\lock.jpg")
    lock_img=lock_img.resize((105,84))
    lock_img=ImageTk.PhotoImage(lock_img)
    
    lock_img_lab=tk.Label(window,image=lock_img)
    lock_img_lab.place(x=0,y=0)

    frame1=tk.Frame(window,width=650,height=420,bg="white")
    frame1.place(x=0,y=80)

    num_lab=tk.Label(frame1,text="Password Length : ",bg='white',font=('veradana bold',14))
    num_lab.place(x=140,y=98)

    num=tk.IntVar()
    length_list=ttk.Combobox(frame1,textvariable=num,width=10,font=("verdana Bold",10))
    length_list["values"]=(1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20)
    length_list.current(5)
    length_list.place(x=310,y=100)

    generate_but=tk.Button(frame1,text="Generate Password",width=40,height=2,fg="dark blue",bg="light blue",borderwidth=2,relief='ridge',justify='center',activeforeground="dark blue",font=("verdana Bold",12),command=lambda:generate(frame1,length_list,window))
    generate_but.place(x=100,y=200)

    saved_pass_but=tk.Button(frame1,text="Saved Passwords",width=40,height=2,fg="dark blue",bg="light blue",borderwidth=2,relief='ridge',justify='center',activeforeground="dark blue",font=("verdana Bold",12),command=lambda:sign_up(frame1,email_entry=tk.Entry(window)))
    saved_pass_but.place(x=100,y=260)


    window.mainloop()


if __name__ == "__main__":
    window=tk.Tk()

    window.configure(bg='white')
    window.title("OTP locker ")
    window.geometry("650x420")
    window.minsize(650,420)
    window.maxsize(650,420)
    window.iconbitmap("Resources\log.ico")

    back_img=Image.open("Resources\\back.jpg")
    back_img=back_img.resize((60,60))
    back_img=ImageTk.PhotoImage(back_img)

    save_img=Image.open("Resources\save.png")
    save_img=save_img.resize((150,60))
    save_img=ImageTk.PhotoImage(save_img)

    bg_image=Image.open("Resources\login_bg.jpg")
    bg_image=bg_image.resize((700,500))
    bg_image=ImageTk.PhotoImage(bg_image)

    login_bg_image=Image.open("Resources\login_bg_img.jpg")
    login_bg_image=login_bg_image.resize((400,300))
    login_bg_image=ImageTk.PhotoImage(login_bg_image)

    user_img=Image.open('Resources\\user.png')
    user_img=user_img.resize((100,100))
    user_img=ImageTk.PhotoImage(user_img)

    email_img=Image.open("Resources\email_img.png")
    email_img=email_img.resize((20,20))
    email_img=ImageTk.PhotoImage(email_img)

    sign_up_img=Image.open("Resources\sign_up.png")
    sign_up_img=sign_up_img.resize((140,40))
    sign_up_img=ImageTk.PhotoImage(sign_up_img)

    submit_img=Image.open("Resources\submit.gif")
    submit_img=submit_img.resize((140,40))
    submit_img=ImageTk.PhotoImage(submit_img)


    file = "Resources\Saved_Passwords.xlsx"
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    if ws["C2"].value==None:
        registration(window)
    else:
        main()
        
    

